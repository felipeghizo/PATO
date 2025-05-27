from flask import Blueprint, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import xml.etree.ElementTree as ET
import logging
import re
from collections import defaultdict
import os
import fitz  # PyMuPDF
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from typing import Dict, Set, Tuple, Optional
import tempfile
import re

# Configure logging
logging.basicConfig(
    filename='comparador.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

easyckd_bp = Blueprint('easyckd', __name__, template_folder='../templates/easy_ckd')

# Global variables to store session data
session_data = {
    'bom_file': None,
    'xml_files': {},
    'pdf_files': {},
    'pcba_positions': None,
    'available_versions': set(),
    'results': {},
    'versions': []
}

class PDFHighlighter:
    def __init__(self):
        self.logger = logging.getLogger('PDFHighlighter')
        self.config = {
            'verified': {'color': (0.2, 0.8, 0.2), 'label': "TOP"},
            'verified_bottom': {'color': (0.6, 0.2, 0.8), 'label': "BOTTOM"},
            'extra': {'color': (0.9, 0.7, 0.1), 'label': "EXTRA"},
            'error': {'color': (1, 0, 0), 'label': "ERRO"}
        }
        self.highlight_opacity = 0.5
        self.font_size = 7
        self.marked_areas = set()
        self.text_position_cache = {}

    def enhance_pdf_marking(self, pdf_path: str, positions_data: Dict[str, Dict], pcba_code: str) -> Optional[str]:
        try:
            extra_positions = [pos for pos, data in positions_data.items() 
                            if data.get('status') == 'extra' or data.get('Top+Bottom') == 'EXTRA']
            logging.info(f"Posições EXTRA a serem marcadas: {extra_positions}")
            
            doc = fitz.open(pdf_path)
            self.marked_areas = set()
            self.text_position_cache = {}
            
            for page_num, page in enumerate(doc):
                self._cache_text_positions(page, page_num)
            
            for page_num, page in enumerate(doc):
                self.logger.info(f"Processando página {page_num + 1}")
                for pos, data in positions_data.items():
                    color = self._determine_marking_color(data)
                    self._mark_position_precise(page, pos, color, page_num)
                self._add_compact_legend(page)
            
            temp_dir = tempfile.gettempdir()
            output_filename = f"{pcba_code}_marked.pdf"
            output_path = os.path.join(temp_dir, output_filename)
            doc.save(output_path, garbage=4, deflate=True, pretty=True)
            
            return output_path if os.path.exists(output_path) else None
            
        except Exception as e:
            self.logger.error(f"Falha crítica: {str(e)}", exc_info=True)
            return None
        finally:
            if 'doc' in locals():
                doc.close()

    def _cache_text_positions(self, page, page_num):
        for word in page.get_text("words"):
            text = word[4].strip()
            if text:
                self.text_position_cache[(page_num, text)] = word[:4]

    def _determine_marking_color(self, data):
        if data.get('status') == 'error':
            return self.config['error']['color']
        elif data.get('status') == 'extra' or data.get('Top+Bottom') == 'EXTRA':
            return self.config['extra']['color']
        elif data.get('side') == 'bottom':
            return self.config['verified_bottom']['color']
        else:
            return self.config['verified']['color']

    def _mark_position_precise(self, page, pos, color, page_num):
        cache_key = (page_num, pos)
        if cache_key in self.text_position_cache:
            rect = self.text_position_cache[cache_key]
            highlight = page.add_highlight_annot(rect)
            highlight.set_colors(stroke=color)
            highlight.set_opacity(self.highlight_opacity)
            highlight.update()
            self.marked_areas.add(pos)

    def _add_compact_legend(self, page):
        legend_x = page.rect.width - 120
        legend_y = page.rect.height - 50
        for key, config in self.config.items():
            page.draw_rect(
                [legend_x, legend_y, 15, 15],
                color=config['color'],
                fill=config['color'],
                overlay=True
            )
            page.insert_text(
                (legend_x + 20, legend_y + 10),
                config['label'],
                fontsize=self.font_size,
                color=(0, 0, 0)
            )
            legend_y += 20

def identify_component_column(df: pd.DataFrame) -> Optional[str]:
    for col in df.columns:
        if df[col].dropna().astype(str).str.fullmatch(r'^\d{7}$').any():
            return col
    return None

def identify_quantity_column(df: pd.DataFrame) -> Optional[str]:
    common_names = ["Real qty", "Quantity", "Qty", "Quantidade"]
    for col in df.columns:
        if any(name.lower() in col.lower() for name in common_names):
            if pd.api.types.is_numeric_dtype(df[col]):
                if (df[col].dropna() % 1 == 0).mean() > 0.95:
                    return col
    return None

def load_bom(bom_file: str) -> Tuple[Optional[Dict], Optional[Set], Optional[str], Optional[str]]:
    try:
        if bom_file.endswith('.xlsx'):
            bom_df = pd.read_excel(bom_file, engine='openpyxl')
        elif bom_file.endswith('.csv'):
            bom_df = pd.read_csv(bom_file, encoding_errors='ignore')
        else:
            raise ValueError("Formato não suportado. Use .xlsx ou .csv")

        component_col = identify_component_column(bom_df)
        quantity_col = identify_quantity_column(bom_df)
        
        if not component_col or not quantity_col or 'Position' not in bom_df.columns:
            raise ValueError("Colunas não identificadas")

        pcba_positions = defaultdict(lambda: {'positions': defaultdict(list), 'components': defaultdict(int)})
        available_versions = set()

        for _, row in bom_df.iterrows():
            position_data = row['Position']
            component_code = row[component_col]
            quantity = row[quantity_col]

            if pd.isna(position_data) or pd.isna(component_code) or pd.isna(quantity):
                continue

            for pcba_entry in str(position_data).split(';'):
                pcba_entry = pcba_entry.strip()
                if not pcba_entry or ':' not in pcba_entry:
                    continue

                try:
                    pcba_code, positions = pcba_entry.split(':', 1)
                    pcba_code = pcba_code.strip()
                    
                    for pos in positions.split(','):
                        pos = pos.strip()
                        if pos:
                            pcba_positions[pcba_code]['positions'][pos] = {'component': component_code}
                    
                    pcba_positions[pcba_code]['components'][component_code] += int(quantity)
                    available_versions.add(pcba_code)
                except ValueError as e:
                    logging.warning(f"Formato inválido na BOM: {pcba_entry} - {str(e)}")

        return dict(pcba_positions), available_versions, component_col, quantity_col

    except Exception as e:
        logging.error(f"Falha ao processar BOM: {str(e)}")
        return None, None, None, None

def load_package_positions(xml_file: str) -> Tuple[Optional[Dict], Optional[Set]]:
    try:
        positions = set()
        side_info = {}
        tree = ET.parse(xml_file)
        
        for element in tree.findall('.//Element'):
            name = element.get('Name', '').strip()
            if not name or name in ['ASS-TOP', 'ASS-BOTTOM']:
                continue
                
            if not any(x in name.upper() for x in ['ASS-TOP', 'ASS-BOTTOM']):
                positions.add(name)
                side_info[name] = 'top'
                
                for subelement in element.findall('Element'):
                    subelement_name = subelement.get('Name', '').strip()
                    if subelement_name == "ASS-BOTTOM":
                        side_info[name] = 'bottom'
                    elif subelement_name == "ASS-TOP":
                        side_info[name] = 'top'
        
        return side_info, positions
        
    except Exception as e:
        logging.error(f"Erro ao processar XML: {str(e)}")
        return None, None

def compare_positions(pcba_positions: Dict, package_data: Tuple[Dict, Set], pcba_code: str) -> Optional[Tuple[pd.DataFrame, Set, Set]]:
    if pcba_code not in pcba_positions:
        return None

    side_info, package_positions = package_data
    bom_data = pcba_positions[pcba_code]
    results = []
    missing_in_package = set()
    missing_in_bom = set(package_positions)

    global_component_counts = defaultdict(int)
    for version, bom_data_all in pcba_positions.items():
        for pos, pos_data in bom_data_all['positions'].items():
            component = pos_data['component']
            global_component_counts[component] += 1

    component_summary = defaultdict(lambda: {
        'qtd_bom': 0,
        'qtd_xml_total': 0,
        'qtd_xml_top': 0,
        'qtd_xml_bottom': 0,
        'lados': set(),
        'posicoes_top': [],
        'posicoes_bottom': [],
        'faltando_top': [],
        'faltando_bottom': []
    })

    for pos, data in bom_data['positions'].items():
        component = data['component']
        side = side_info.get(pos, 'top')
        
        comp_data = component_summary[component]
        comp_data['qtd_bom'] = bom_data['components'].get(component, 0)
        comp_data['lados'].add(side)

        if pos in package_positions:
            xml_side = side_info.get(pos, 'top')
            if xml_side == 'top':
                comp_data['qtd_xml_top'] += 1
                comp_data['posicoes_top'].append(pos)
            else:
                comp_data['qtd_xml_bottom'] += 1
                comp_data['posicoes_bottom'].append(pos)
            comp_data['qtd_xml_total'] = comp_data['qtd_xml_top'] + comp_data['qtd_xml_bottom']
        else:
            if side == 'top':
                comp_data['faltando_top'].append(pos)
            else:
                comp_data['faltando_bottom'].append(pos)
            missing_in_package.add(pos)

        if pos in missing_in_bom:
            missing_in_bom.remove(pos)

    for component, data in component_summary.items():
        total_global = global_component_counts.get(component, 0)
        verificacao_top_bottom = f"GLOBAL: {total_global}"
        
        if data['qtd_bom'] != total_global:
            verificacao_top_bottom += " (ERRO)"

        if 'top' in data['lados'] and 'bottom' in data['lados']:
            results.append({
                'Componente': component,
                'Qtd BOM': data['qtd_bom'],
                'Qtd XML (Top)': data['qtd_xml_top'],
                'Qtd XML (Bottom)': 0,
                'Top+Bottom': verificacao_top_bottom,
                'Status': "OK" if data['qtd_xml_top'] > 0 else "FALTANDO",
                'Lado': 'TOP',
                'Posições': ', '.join(sorted(data['posicoes_top'])) if data['posicoes_top'] else '-',
                'Faltando': ', '.join(sorted(data['faltando_top'])) if data['faltando_top'] else '-'
            })
            
            results.append({
                'Componente': component,
                'Qtd BOM': data['qtd_bom'],
                'Qtd XML (Top)': 0,
                'Qtd XML (Bottom)': data['qtd_xml_bottom'],
                'Top+Bottom': verificacao_top_bottom,
                'Status': "OK" if data['qtd_xml_bottom'] > 0 else "FALTANDO",
                'Lado': 'BOTTOM',
                'Posições': ', '.join(sorted(data['posicoes_bottom'])) if data['posicoes_bottom'] else '-',
                'Faltando': ', '.join(sorted(data['faltando_bottom'])) if data['faltando_bottom'] else '-'
            })
        else:
            lado = 'TOP' if 'top' in data['lados'] else 'BOTTOM'
            qtd_xml = data['qtd_xml_top'] if lado == 'TOP' else data['qtd_xml_bottom']
            posicoes = data['posicoes_top'] if lado == 'TOP' else data['posicoes_bottom']
            faltando = data['faltando_top'] if lado == 'TOP' else data['faltando_bottom']
            
            results.append({
                'Componente': component,
                'Qtd BOM': data['qtd_bom'],
                'Qtd XML (Top)': data['qtd_xml_top'] if lado == 'TOP' else 0,
                'Qtd XML (Bottom)': data['qtd_xml_bottom'] if lado == 'BOTTOM' else 0,
                'Top+Bottom': verificacao_top_bottom,
                'Status': "OK" if qtd_xml > 0 else "FALTANDO",
                'Lado': lado,
                'Posições': ', '.join(sorted(posicoes)) if posicoes else '-',
                'Faltando': ', '.join(sorted(faltando)) if faltando else '-'
            })

    result_df = pd.DataFrame(results)
    if not result_df.empty:
        result_df = result_df.sort_values('Lado', ascending=False)

    return result_df, missing_in_bom, missing_in_package

@easyckd_bp.route('/upload_bom', methods=['POST'])
def upload_bom():
    try:
        if 'bom_file' not in request.files:
            return jsonify({'status': 'error', 'message': 'Nenhum arquivo enviado'}), 400

        file = request.files['bom_file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'Nenhum arquivo selecionado'}), 400

        filename = secure_filename(file.filename)
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        file.save(filepath)
        
        pcba_positions, available_versions, comp_col, qty_col = load_bom(filepath)
        if pcba_positions:
            session_data['bom_file'] = filepath
            session_data['pcba_positions'] = pcba_positions
            session_data['available_versions'] = available_versions
            
            return jsonify({
                'status': 'success',
                'message': f"BOM carregada com sucesso!\nColuna Componente: {comp_col}\nColuna Quantidade: {qty_col}\nVersões disponíveis: {', '.join(available_versions)}",
                'versions': list(available_versions)
            })
        
        return jsonify({'status': 'error', 'message': 'Falha ao processar BOM'}), 400
    
    except Exception as e:
        logging.error(f"Erro no upload da BOM: {str(e)}")
        return jsonify({
            'status': 'error', 
            'message': f'Erro interno ao processar BOM: {str(e)}'
        }), 500

@easyckd_bp.route('/upload_xml', methods=['POST'])
def upload_xml():
    try:
        if 'xml_file' not in request.files:
            return jsonify({
                'status': 'error',
                'message': 'Nenhum arquivo XML enviado'
            }), 400

        file = request.files['xml_file']
        version = request.form.get('version')
        
        if not version:
            return jsonify({
                'status': 'error',
                'message': 'Versão não especificada'
            }), 400
            
        if file.filename == '':
            return jsonify({
                'status': 'error', 
                'message': 'Nenhum arquivo selecionado'
            }), 400

        filename = secure_filename(file.filename)
        temp_dir = tempfile.gettempdir()
        filepath = os.path.join(temp_dir, filename)
        file.save(filepath)
        
        # Validate it's actually an XML file
        try:
            ET.parse(filepath)  # Try parsing to validate XML
        except ET.ParseError:
            os.remove(filepath)
            return jsonify({
                'status': 'error',
                'message': 'Arquivo não é um XML válido'
            }), 400
        
        session_data['xml_files'][version] = filepath
        return jsonify({
            'status': 'success',
            'message': f"XML carregado para versão {version}",
            'version': version,
            'filename': filename
        })
        
    except Exception as e:
        logging.error(f"Erro no upload do XML: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Erro interno ao processar XML: {str(e)}'
        }), 500

@easyckd_bp.route('/upload_pdf', methods=['POST'])
def upload_pdf():
    if 'pdf_file' not in request.files:
        return jsonify({'status': 'error', 'message': 'Nenhum arquivo PDF enviado'}), 400

    file = request.files['pdf_file']
    version = request.form.get('version')
    
    if not version:
        return jsonify({'status': 'error', 'message': 'Versão não especificada'}), 400
        
    if file.filename == '':
        return jsonify({'status': 'error', 'message': 'Nenhum arquivo selecionado'}), 400

    filename = secure_filename(file.filename)
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    file.save(filepath)
    
    session_data['pdf_files'][version] = filepath
    return jsonify({
        'status': 'success',
        'message': f"PDF carregado para versão {version}",
        'version': version,
        'filename': filename
    })

@easyckd_bp.route('/compare', methods=['POST'])
def compare():
    try:
        # Validate required data exists
        if not session_data.get('bom_file') or not session_data.get('xml_files'):
            return jsonify({
                'status': 'error',
                'message': 'Carregue a BOM e os XMLs primeiro'
            }), 400
        
        session_data['results'] = {}
        output = []
        
        for version, xml_file in session_data['xml_files'].items():
            try:
                # Load and validate XML
                side_info, positions = load_package_positions(xml_file)
                if not positions:
                    output.append(f"AVISO: Nenhuma posição encontrada no XML para {version}")
                    continue
                
                # Compare with BOM data
                result = compare_positions(
                    session_data['pcba_positions'],
                    (side_info, positions),
                    version
                )
                
                if not result:
                    output.append(f"AVISO: Nenhum resultado para versão {version}")
                    continue
                
                result_df, missing_bom, missing_pkg = result
                session_data['results'][version] = (result_df, missing_bom, missing_pkg)
                
                # Format results
                output.append(f"\n=== RESULTADOS PARA VERSÃO {version} ===")
                output.append(result_df.to_string(index=False, justify='center'))
                
                if missing_pkg:
                    output.append(f"\nPosições faltando no XML ({len(missing_pkg)}):")
                    output.append(", ".join(sorted(missing_pkg)))
                
                if missing_bom:
                    output.append(f"\nPosições extras no XML ({len(missing_bom)}):")
                    output.append(", ".join(sorted(missing_bom)))
                
                output.append("\n" + "="*80 + "\n")
            
            except Exception as e:
                error_msg = f"ERRO ao processar {version}: {str(e)}"
                output.append(error_msg)
                logging.error(error_msg, exc_info=True)
        
        return jsonify({
            'status': 'success',
            'message': '\n'.join(output),
            'versions': list(session_data['results'].keys())
        })
    
    except Exception as e:
        error_msg = f"Erro geral na comparação: {str(e)}"
        logging.error(error_msg, exc_info=True)
        return jsonify({
            'status': 'error',
            'message': error_msg
        }), 500

@easyckd_bp.route('/export_excel', methods=['POST'])
def export_excel():
    try:
        if not session_data.get('results'):
            return jsonify({
                'status': 'error',
                'message': 'Nenhum resultado disponível para exportação'
            }), 400

        temp_dir = tempfile.gettempdir()
        filename = f"Relatorio_EasyCKD_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for version, (result_df, missing_bom, missing_pkg) in session_data['results'].items():
                    # Limit sheet name to 31 chars (Excel limitation)
                    sheet_name = version[:25] + '..' if len(version) > 25 else version
                    result_df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Add discrepancies sheet
                    discrep_df = pd.DataFrame({
                        'Tipo': ['Faltando no XML', 'Extra no XML'],
                        'Quantidade': [len(missing_pkg), len(missing_bom)],
                        'Posições': [', '.join(sorted(missing_pkg)), ', '.join(sorted(missing_bom))]
                    })
                    discrep_sheet_name = f"{sheet_name[:20]}_DISCREP" if len(sheet_name) > 20 else f"{sheet_name}_DISCREP"
                    discrep_df.to_excel(writer, sheet_name=discrep_sheet_name, index=False)
            
            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            logging.error(f"Erro ao gerar Excel: {str(e)}")
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({
                'status': 'error',
                'message': f'Falha ao gerar arquivo Excel: {str(e)}'
            }), 500
            
    except Exception as e:
        logging.error(f"Erro no export_excel: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Erro interno ao exportar Excel: {str(e)}'
        }), 500

@easyckd_bp.route('/export_pdf', methods=['POST'])
def export_pdf():
    try:
        version = request.form.get('version')
        if not version or not session_data.get('results') or not session_data.get('pdf_files'):
            return jsonify({
                'status': 'error',
                'message': 'Dados incompletos para exportação'
            }), 400

        if version not in session_data['results'] or version not in session_data['pdf_files']:
            return jsonify({
                'status': 'error',
                'message': f'Versão {version} não encontrada nos resultados'
            }), 404

        try:
            result_df, missing_in_bom, missing_in_pkg = session_data['results'][version]
            pdf_file = session_data['pdf_files'][version]
            
            # Process positions data
            side_info, _ = load_package_positions(session_data['xml_files'][version])
            positions_data = {}
            
            for pos in session_data['pcba_positions'][version]['positions']:
                status_row = result_df[
                    (result_df['Posições'].str.contains(rf'\b{pos}\b', regex=True, na=False)) |
                    (result_df['Faltando'].str.contains(rf'\b{pos}\b', regex=True, na=False))
                ]
                
                if not status_row.empty:
                    status = status_row['Status'].values[0]
                    top_bottom = status_row['Top+Bottom'].values[0]
                    side = side_info.get(pos, 'top')
                    
                    is_error = (status == "FALTANDO" or '(ERRO)' in str(top_bottom) or (pos in missing_in_pkg))
                    is_extra = (pos in missing_in_bom)
                    
                    positions_data[pos] = {
                        'status': 'error' if is_error else ('extra' if is_extra else 'ok'),
                        'side': side,
                        'Top+Bottom': top_bottom
                    }
            
            # Mark extra positions
            for pos in missing_in_bom:
                if pos:
                    positions_data[pos] = {'status': 'extra', 'Top+Bottom': 'EXTRA', 'side': 'none'}
            
            pdf_highlighter = PDFHighlighter()
            output_path = pdf_highlighter.enhance_pdf_marking(pdf_file, positions_data, version)
            
            if output_path and os.path.exists(output_path):
                return send_file(
                    output_path,
                    as_attachment=True,
                    download_name=f"{version}_marked.pdf",
                    mimetype='application/pdf'
                )
            else:
                return jsonify({
                    'status': 'error',
                    'message': 'Falha ao gerar PDF marcado'
                }), 500
                
        except Exception as e:
            logging.error(f"Erro ao processar PDF: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'Erro ao marcar PDF: {str(e)}'
            }), 500
            
    except Exception as e:
        logging.error(f"Erro no export_pdf: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Erro interno ao exportar PDF: {str(e)}'
        }), 500

@easyckd_bp.route('/export_sap', methods=['POST'])
def export_sap():
    try:
        if not session_data.get('results'):
            return jsonify({
                'status': 'error', 
                'message': 'Nenhum resultado disponível para exportação'
            }), 400

        DEFAULT_VALUES = {
            'Component UoM': 'PEÇ',
            'Válido desde': '01.01.1900',
            'Válido até': '31.12.9999',
            'Co-produto': 'Não',
            'Item dummy': 'Não',
            'Recursividade permitida': 'Não',
            'Relevância para cálculo de custos': 'X',
            'Relev.p/produção': 'Sim',
            'Seqüência hierárq.': '0',
            'Probabilidade de utilização (%)': '0',
            'Categoria do item': 'L'
        }

        HEADERS = [
            'Nº item', 'Componente', 'Denominação do componente', 
            'Quantidade do componente', 'Quantidade de posições no lado',
            'Component UoM', 'Texto do item', 'Válido desde', 'Válido até',
            'Nº da modificação', 'Co-produto', 'Item dummy',
            'Recursividade permitida', 'Relevância para cálculo de custos',
            'Relev.p/produção', 'Grupo de itens alternativos',
            'Seqüência hierárq.', 'Probabilidade de utilização (%)',
            'Categoria do item', 'Item Category Description',
            'Criado por', 'Data de criação', 'Modificado por'
        ]

        temp_dir = tempfile.gettempdir()
        filename = f"Relatorio_SAP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        try:
            wb = Workbook()
            wb.remove(wb.active)

            for version, (result_df, _, _) in session_data['results'].items():
                if result_df.empty:  # Pula versões sem dados
                    continue
                    
                safe_version = re.sub(r'[^\w]', '_', version)[:25]
                
                for side in ['TOP', 'BOTTOM']:
                    ws = wb.create_sheet(f"{safe_version}_{side}")
                    ws.append(HEADERS)
                    
                    # Formatação do cabeçalho
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                    
                    side_data = result_df[result_df['Lado'] == side]
                    if side_data.empty:  # Pula lados sem dados
                        continue
                        
                    item_counter = 10
                    
                    for _, row in side_data.iterrows():
                        positions = re.sub(r'\s+', '', str(row['Posições'])) if row['Posições'] else ''
                        if not positions:
                            continue
                            
                        # Divide as posições em chunks
                        pos_list = [p.strip() for p in positions.split(',') if p.strip()]
                        chunks = []
                        current_chunk = []
                        current_length = 0
                        
                        for pos in pos_list:
                            required_length = len(pos) + (1 if current_chunk else 0)
                            if current_length + required_length > 40 and current_chunk:
                                chunks.append((','.join(current_chunk), len(current_chunk)))
                                current_chunk = [pos]
                                current_length = len(pos)
                            else:
                                current_chunk.append(pos)
                                current_length += required_length
                        
                        if current_chunk:
                            chunks.append((','.join(current_chunk), len(current_chunk)))
                        
                        xml_quantity = row['Qtd XML (Top)'] if side == 'TOP' else row['Qtd XML (Bottom)']
                        
                        for chunk, pos_count in chunks:
                            row_data = [
                                f"{item_counter:04d}",
                                str(row['Componente']),
                                '',  # Denominação do componente
                                pos_count,
                                xml_quantity,
                                DEFAULT_VALUES['Component UoM'],
                                chunk,
                                DEFAULT_VALUES['Válido desde'],
                                DEFAULT_VALUES['Válido até'],
                                '',  # Nº da modificação
                                *[DEFAULT_VALUES[k] for k in list(DEFAULT_VALUES.keys())[3:]]
                            ]
                            
                            ws.append(row_data)
                            ws.cell(row=ws.max_row, column=5).fill = PatternFill(
                                start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            
                            item_counter += 10
            
            wb.save(filepath)
            
            if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
                return send_file(
                    filepath,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            return jsonify({
                'status': 'error',
                'message': 'Falha ao gerar arquivo SAP (arquivo vazio ou não criado)'
            }), 500
                
        except Exception as e:
            logging.error(f"Erro ao gerar SAP: {str(e)}", exc_info=True)
            if 'filepath' in locals() and os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({
                'status': 'error',
                'message': f'Falha ao gerar arquivo SAP: {str(e)}'
            }), 500
            
    except Exception as e:
        logging.error(f"Erro no export_sap: {str(e)}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Erro interno ao exportar SAP: {str(e)}'
        }), 500

@easyckd_bp.route('/reset', methods=['POST'])
def reset():
    session_data['bom_file'] = None
    session_data['xml_files'] = {}
    session_data['pdf_files'] = {}
    session_data['pcba_positions'] = None
    session_data['available_versions'] = set()
    session_data['results'] = {}
    session_data['versions'] = []
    return jsonify({'status': 'success', 'message': 'Reset completo realizado'})

@easyckd_bp.route('/', methods=['GET'])
def easyCKD_home():
    return render_template('easy_ckd/easyCKD.html')