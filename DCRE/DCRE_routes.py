from flask import Blueprint, request, render_template, send_file
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
import pandas as pd
import os
import shutil
import logging

# Configura logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

dcre_bp = Blueprint('dcre', __name__, template_folder='../templates/dcre')

@dcre_bp.route('/')
def dcre_home():
    return render_template('dcre/dcre.html')

@dcre_bp.route('/DCRE_Tool', methods=['GET', 'POST'])
def DCRE_Tool():
    logger.debug("Iniciando processamento DCRE")

    if 'excel_file' not in request.files:
        error_message = "Nenhum arquivo encontrado."
        logger.error(error_message)
        return render_template('dcre/dcre_error.html', error_message=error_message)

    file = request.files['excel_file']
    logger.debug(f"Arquivo recebido: {file.filename}")

    if file.filename == '':
        error_message = "Por favor, selecione o arquivo da estrutura do produto."
        logger.error(error_message)
        return render_template('dcre/dcre_error.html', error_message=error_message)

    try:
        # Lê o arquivo sem cabeçalhos
        df = pd.read_excel(file, header=None)
        df.columns = range(df.shape[1])  # Colunas 0 a 7
        logger.debug("DataFrame criado com sucesso do Excel")
        
        # DEBUG: Mostra a estrutura do arquivo
        logger.debug("Estrutura do arquivo:")
        with pd.option_context('display.max_rows', 20, 'display.max_columns', None):
            logger.debug(df.head(20))
    except Exception as e:
        logger.exception("Erro ao ler arquivo XLSX")
        error_message = "Erro ao processar o arquivo XLSX. Verifique se está no padrão P&D exportado pelo SAP."
        return render_template('dcre/dcre_error.html', error_message=error_message)

    if df.empty or df.shape[1] < 8:
        error_message = "Arquivo vazio ou estrutura inválida."
        logger.error(error_message)
        return render_template('dcre/dcre_error.html', error_message=error_message)

    # ENCONTRA A LINHA INICIAL DOS DADOS
    start_row = 0
    for i in range(len(df)):
        if str(df.iloc[i,0]).strip().isdigit():  # Procura a primeira linha onde a coluna A tem um número
            start_row = i
            break

    # Função para limpar códigos
    def limpar_codigo(valor):
        try:
            return int(float(str(valor).strip()))
        except (ValueError, TypeError):
            return None

    # Aplica à coluna D (Nº componente) - índice 3
    # Ajusta para usar start_row como offset
    df_componentes = df.iloc[start_row:]
    df_componentes[3] = df_componentes[3].apply(limpar_codigo)
    df_componentes = df_componentes[df_componentes[3].notnull()]

    # Converte quantidades (coluna G - índice 6)
    df_componentes[6] = df_componentes[6].astype(str).str.replace(',', '.')

    # Processa nível de explosão (coluna A - índice 0)
    df_componentes[0] = df_componentes[0].astype(str).str.replace(".", "", regex=False)
    df_componentes[0] = pd.to_numeric(df_componentes[0], errors='coerce')

    # Verifica estrutura completa
    if (not ((df_componentes[0] > 1).any()) and ((df_componentes[3] > 2000000) & (df_componentes[3] < 3000000)).any()):
        error_message = "Este arquivo não contempla todos os níveis da estrutura, verifique novamente."
        return render_template('dcre/dcre_error.html', error_message=error_message)

    try:
        ncm = request.form['ncm']
        peso = request.form['weight']
        modelo = request.form['model']
        preco = request.form['price']
    except KeyError as e:
        error_message = f"Campo faltando: {str(e)}"
        logger.error(error_message)
        return render_template('dcre/dcre_error.html', error_message=error_message)

    # Renomeia colunas para agrupamento
    dff = df_componentes.rename(columns={
        3: 'Num_componente',       # Coluna D
        4: 'Texto_breve_objeto',   # Coluna E
        7: 'Unid_med_componente',  # Coluna H
        6: 'Qtd_componente_UMC'    # Coluna G
    })

    # Agrupamento e filtros
    dff["Qtd_componente_UMC"] = pd.to_numeric(dff["Qtd_componente_UMC"], errors='coerce')
    dff = dff.groupby(["Num_componente", "Texto_breve_objeto", "Unid_med_componente"], as_index=False).agg({"Qtd_componente_UMC": "sum"})
    
    dfn = dff.query('(Num_componente < 1870000 or (Num_componente >= 1880000 and Num_componente < 2000000))')
    dfi = dff.query('(Num_componente >= 1870000 and Num_componente < 1880000) or (Num_componente >= 1820000 and Num_componente < 1830000) or (Num_componente >= 3000000 and Num_componente < 4000000)')

    # Conversão para listas
    codigo_nacional = dfn["Num_componente"].tolist()
    quantidade_nacional = [str(x).replace('.', ',') for x in dfn["Qtd_componente_UMC"].tolist()]
    descricao_nacional = dfn["Texto_breve_objeto"].tolist()
    unidade_nacional = dfn["Unid_med_componente"].tolist()

    codigo_internacional = dfi["Num_componente"].tolist()
    quantidade_internacional = [str(x).replace('.', ',') for x in dfi["Qtd_componente_UMC"].tolist()]
    descricao_internacional = dfi["Texto_breve_objeto"].tolist()
    unidade_internacional = dfi["Unid_med_componente"].tolist()

    # Carrega template do Excel
    base_dir = os.path.dirname(__file__)
    template_path = os.path.join(base_dir, 'Template_DCRE.xlsx')

    # Obtém o código do produto (linha 0 na imagem = start_row no DataFrame)
    try:
        codigo_produto = int(df.iloc[start_row, 1])  # Material
        descricao = str(df.iloc[start_row+2, 1]).strip()  # Descrição (3 linhas abaixo do Material)
        centr_util = str(df.iloc[start_row+1, 1]).strip()  # Centr/Util./Alt.
    except Exception as e:
        logger.exception("Erro ao obter dados do produto")
        error_message = "Erro ao processar os dados principais do produto."
        return render_template('dcre/dcre_error.html', error_message=error_message)

    # Define nome do arquivo final
    output_filename = f"DCRE-{codigo_produto}.xlsx"
    output_path = os.path.join(base_dir, output_filename)

    # Cria cópia do template
    shutil.copy(template_path, output_path)

    # Carrega a cópia para modificar
    workbook = load_workbook(output_path)

    # Preenche dados principais
    sheet_dcre = workbook["Folha de Rosto - DCRE"]
    sheet_dcre["B2"] = codigo_produto
    sheet_dcre["B4"] = descricao
    sheet_dcre["B11"] = f"{peso.replace('.', ',')} kg"
    sheet_dcre["B3"] = modelo
    sheet_dcre["B5"] = f"R$ {preco.replace('.', ',')}"
    sheet_dcre["B6"] = ncm

    # Define estilo de borda
    border = Border(
        left=Side(border_style="thin", color='00000000'),
        right=Side(border_style="thin", color='00000000'),
        top=Side(border_style="thin", color='00000000'),
        bottom=Side(border_style="thin", color='00000000')
    )

    # Função para preencher abas
    def preencher_aba(sheet, codigos, quantidades, descricoes, unidades, config):
        for i in range(len(codigos)):
            row = i + 2
            sheet.cell(row=row, column=config["codigo"]).value = codigos[i]
            sheet.cell(row=row, column=config["descricao"]).value = descricoes[i]
            sheet.cell(row=row, column=config["vazio"]).value = "vazio"
            sheet.cell(row=row, column=config["quantidade"]).value = quantidades[i]
            sheet.cell(row=row, column=config["quantidade"]).number_format = '0.00000'
            sheet.cell(row=row, column=config["unidade"]).value = unidades[i]

            for key in config:
                sheet.cell(row=row, column=config[key]).border = border

            for col in range(1, 12):
                if col not in config.values():
                    sheet.cell(row=row, column=col).value = ""
                    sheet.cell(row=row, column=col).border = border

    # Configurações das abas
    config_importado = {
        "codigo": 3,
        "descricao": 4,
        "vazio": 5,
        "quantidade": 6,
        "unidade": 7
    }

    config_nacional = {
        "codigo": 4,
        "descricao": 5,
        "vazio": 6,
        "quantidade": 7,
        "unidade": 8
    }

    # Preenche abas
    preencher_aba(workbook["Importado"], codigo_internacional, quantidade_internacional, descricao_internacional, unidade_internacional, config_importado)
    preencher_aba(workbook["Nacional"], codigo_nacional, quantidade_nacional, descricao_nacional, unidade_nacional, config_nacional)

    # Salva o arquivo
    workbook.save(output_path)

    # Retorna para download
    return send_file(output_path, as_attachment=True, download_name=output_filename)